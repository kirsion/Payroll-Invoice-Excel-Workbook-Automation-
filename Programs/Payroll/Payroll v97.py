# Instructions: Drop the ADP invoice report in the 'Raw Payroll Invoices' folder, file name doesn't matter.
# Multiple files can be added, but only one temp file with the Excel formula will appear.
# When first running the program, there might be an error. Run the program a second time, the error should disappear.
# Make sure to close out all other Excel files before running the program since the program will force-close any open Excel files.

# v24: Changed the name on the Vision-Post tax
# v26: Removed YUB from sheet title.
# v27: Combined Payroll and Subtotal programs, no more need of separate Subtotal file.
# v32: Added feature that now requires to add payroll invoices to 'Raw Payroll Invoices' folder, and the correctly named file will be created in the root folder,
#      including the raw sheet. Instead of copying and pasting and needing the Excel files already present in the root folder.
# v33: Added option to run smaller adjustment invoices. The new invoice file causes Excel to stay open so added function to kill Excel.exe.
# v34: For attribute error win32com.gen_py 'CLSIDToPackageMap' error. This fix works after restart.
#      In PowerShell, enter: Remove-Item -path $env:LOCALAPPDATA\Temp\gen_py -recurse. If still getting this error, rerun the program.
# v35: Removed the TOTAL ADJ Subtotal column.
# v38: Added the formula sheet back into the workbook and deleted data only sheet.
# v40: Added automated powershell script execution exception for attribute error.
# v42: Added sorting and sums to ETL-1.
# v44: Removed all Excel vlookup and linking, constructed and inputted data in python. Bolding of sums.
# v45: Added exception to pass an error to organize the ETL-2 sheet if the yardi code column doesn't exist.
#      Fixed ADJ GL-Code not being written from the previous version.
# v46: Added HACLA sheet.
# v48: Added the CBT-CVS sheet, which is the Transfer sheet, sorted.
# v49: Sorted the Transfer sheet by property name instead of Code.
# v50: Skip the Hacla separate sheet if there are no Hacla rows.
# v51: Make sure that all rows in "Department Number" have a value. Changes any empty/"None" cell to Truncated Home Department #.
# v52: Change the path of "codes" file outside of the HR folder so others can access the file.
# v54: Change so that invoices without any CORP rows will work.
# v55: Made it so the temp file in the Raw Payroll Invoices Folder does not get removed as the temp file has the useful Excel cell formulas.
# v56: Change so the invoice type in the sheet name is dynamic, not static at "P74".
# v57: Added "ADJ 2-UNION" column for RD4 Invoices.
#      Tip: When adding new column, put column between at the ends of columns being used in the locate function.
# v58: Added fixed delivery fee for RD4 invoices only.
# v59: Update Delivery fee column with Ancillary Product Fee.
# v60: Updated with dynamic delivery fee.
# v62: Added extra exception to catch error when codes-current.xlsx is missing new properties. When sorting function fails if column has None values in it.
#      And saves worksheet before error occurs to more easily locate were the None error is.
# v63: Changed "Department Number" column to "Cost Number".
# v64: Added win32com exception for "win32.gencache.EnsureDispatch" to "win32.Dispatch" if EnsureDispatch causes an error.
# v65: Added another function that deletes gen.py in Temp folder which may cause 'CLSIDToPackageMap' error.
# v66: Added separating some 'CORP03' rows that are supposed to be for HACLA.
# v67: Added ADJ-3 UNION IN FEE to columns.
# v68: Added so that if the invoice lacks ADJ-3 UNION IN FEE column, a column for it will be automatically added.
# v69: Changed "Home Department Description" to "Cost Number Description".
#      When downloading the custom report, make sure to checkmark "min" or "max" in the Selected fields for "Cost Number Description".
# v70: Added so that sorting error exception outputs which properties that need to be added to the codes-current Excel file.
# v71: Change Cost Number has a value check, instead of adding CORP00, it notifies the user to review invoice.
# v72: Updated if the raw invoice has mistakes, such as P74 or missing cell values, it will display which is missing and end the program.
# v73: Changed "Cost Number Description" to "Cost Number Description" so the column is not deleted.
# v74: Sorted and added formulas to the Transfer sheet.
# v75: Skips Transfer Sheet code for RD4 invoice.
# v76: Copies temp file that has Excel formulas sheet and put it into the main workbook. And changed max_col to hard-coded length.
# v77: Made final P74 by copying.
# v79: Hide columns for Final P74.
# v80: Fixed Final P74 sheet by using the non-formula sheet.
# v81: Fixed order list by adding 6911 again.
# v83: Completed the writing, formatting, styling on the Final P74 sheet.
# v84: Copied P74 values onto the Labor Distribution template.
# v87: Fixed some of the exception errors in the raw invoice check.
# v88: Added corp value to the labor distribution.
# v89: Updated the HACLA excluded name list.
# v90: Updated the locate function so there is only one function instead of four different functions. Added default argument values to reduce the number of arguments in function calls.
# v93: Fixed skipping RD4 invoice.
# v94: Fixed Invoice dates and added the Hyder invoice sheet. Fixed Invoice linking codes.
# v95: Updated so that if invoice is missing "INV" or "INV-CHULA" row, a dummy row will be added.
# v96: Add condition for "INV" or "INV-CHULA" row, so that it will skip on RD4.
# v97: Delete "INV" or "INV-CHULA" rows after Subtotal part is complete, except from sheets that have formulas. Removed HACLA CORP03 renaming.

import openpyxl as xl
import pandas as pd
from datetime import timedelta, datetime
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import re
import sys
from string import ascii_uppercase
import itertools
import win32com.client as win32
from win32com.client import Dispatch
from glob import glob as glob
from os import getcwd, chdir
import psutil
import subprocess
from copy import copy
import time


win32c = win32.constants
start_time = time.time()


# This locate function finds any cell in the sheet with a particular string value
# and returns a list of its position information, [row #, column #, and coordinate].
# If there are multiples of the same string, the triplets get added to the end. If information of multiple of the same
# string is needed, then append may be more useful than extend to return a list of list.
# If the string does not exist, then the function returns an empty list.

# Locate function assumes that the max number of columns is 45 in the worksheet. If more columns are added, increase this number.
def locate(str_val, sheet=None, end_col=45, start_col=1):
    if sheet is None:
        sheet = ws
    position_info = []
    for row_s in sheet.iter_rows(min_col=start_col, max_col=end_col):
        for cell_s in row_s:
            if cell_s.value == str_val:
                position_info.extend([cell_s.row, cell_s.column, cell_s.coordinate])
    return position_info if position_info else []


# This function removes numerical characters (and whitespace) from a string of coordinate.
# Since this function calls another function, it is very inefficient when called inside loops.
def col_letter(string):
    return "".join(filter(lambda p_1: p_1.isalpha(), locate(string)[2]))


# Removes all columns besides the Code and Total columns.
def del_col(abs_column):
    raw_column = [ws.cell(row=1, column=k_).value for k_ in range(1, ws.max_column + 1)]
    abs_column_idx = [raw_column.index(col) + 1 for col in abs_column if col in raw_column]
    indexes_set_diff = set(range(1, ws.max_column + 1)).difference(abs_column_idx)
    for w_ in sorted(indexes_set_diff, reverse=True):
        ws.delete_cols(w_, 1)


def del_corp():
    corporate_list = []
    sub_string = ['CORP', '0COR', 'OCOR']

    # Find the CORP rows and store their row indices in corp_list
    for corp_rows in range(1, ws.max_row + 1):  # Start from 1 instead of 0
        cell_value = ws.cell(row=corp_rows, column=1).value
        if cell_value and any(sub in cell_value for sub in sub_string):
            corp_list.append(corp_rows)

    # Determine rows to delete
    rows_to_delete = []
    for del_rows in corporate_list:
        rows_to_delete.extend([del_rows - 2, del_rows - 1, del_rows, del_rows + 1, del_rows + 2])

    # Ensure uniqueness and sort the rows to delete
    rows_to_delete = sorted(set(rows_to_delete))

    # Delete rows from the Excel sheet
    for row_indexes in reversed(rows_to_delete):
        ws.delete_rows(row_indexes, 1)


def length():
    field_total_idx_ = []
    # Finds the length of rows of FIELD data or non-CORP rows.
    s = -1
    while True:
        s += 1
        field_total_idx_.append(ws.cell(row=s + 2, column=1).value)
        if None in field_total_idx_:
            break

    field_total_idx_ = list(filter(None, field_total_idx_))
    # Total of field rows, the ending row number field_row + 1, since the title row.
    length.f_row = len(field_total_idx_)

    return len(field_total_idx_)


empty = Side()
thin = Side(border_style="thin", color="000000")
double = Side(border_style='double')
red = Font(size=10, color='FF0000')


def negative_red(cell_range):
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=red))


def set_font(cell_range):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.font = Font(name="Arial", size=10)


def set_bold(cell_range):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.font = Font(name="Arial", size=10, bold=True)


def set_underline(cell_range):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.font = Font(name="Arial", size=10, bold=True, underline="single")


def no_fill(cell_range):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.fill = PatternFill()


def set_no_border(cell_range):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.border = Border(top=empty, left=empty, right=empty, bottom=empty)


def set_border(cell_range, top_style, left_style, right_style, bottom_style):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.border = Border(top=top_style, left=left_style, right=right_style, bottom=bottom_style)


def set_alignment(cell_range, position):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.alignment = Alignment(horizontal=position)


def fill_color(cell_range, hex_code):
    for rows in ws[cell_range]:
        for cells in rows:
            cells.fill = PatternFill(start_color=hex_code, end_color=hex_code, fill_type='solid')


def manual_adjust(letter, adj_columns, sheet=None):
    if sheet is None:
        sheet = ws

    def all_alphabet():
        for size in itertools.count(1):
            for sizes in itertools.product(ascii_uppercase, repeat=size):
                yield "".join(sizes)

    alphabet_list = []
    for alphabets in all_alphabet():
        alphabet_list.append(alphabets)
        if alphabets == letter:
            break

    col_tup_list = tuple(zip(alphabet_list, adj_columns))
    for col_tup in col_tup_list:
        sheet.column_dimensions[col_tup[0]].width = col_tup[1]


def move_sheet(from_loc=None, to_loc=None):
    # noinspection PyProtectedMember
    sheets = wb._sheets
    # Choose which sheet to move.
    if from_loc is None:
        from_loc = len(sheets) - 1
    # Choose where the sheet will be moved to.
    if to_loc is None:
        to_loc = 0
    sheet = sheets.pop(from_loc)
    sheets.insert(to_loc, sheet)


def kill_excel():
    for proc in psutil.process_iter():
        if proc.name() == "EXCEL.EXE":
            proc.kill()


def sort_excel_sheet(starting_row, starting_column):
    data = []
    for row_1 in ws.iter_rows(min_row=starting_row, values_only=True):
        data.append(row_1)

    # Convert values to string so the comparison is always the same type.
    # Specifically here, only converted the fourth element which is the ABA column.
    str_data = []
    for tup in data:
        new_tuple = []
        for i_2, elem in enumerate(tup):
            if i_2 == 3 and elem is not None:
                new_tuple.append(str(elem))
            else:
                new_tuple.append(elem)
        str_data.append(tuple(new_tuple))

    sorted_data = sorted(str_data, key=lambda x: x[starting_column - 1])

    # Rearrange the rows based on the sorted data
    for i_1, row_1 in enumerate(sorted_data, start=starting_row):
        for j_1, value_1 in enumerate(row_1, start=1):
            ws.cell(row=i_1, column=j_1, value=value_1)


def copy_excel_file(input_file_path, new_file_name):
    # Extract the directory path and file name from the input file path
    directory = getcwd()
    filename = os.path.basename(input_file_path)

    # Move one level up in the directory hierarchy
    parent_directory = os.path.abspath(os.path.join(directory, os.pardir))

    # Create the new file path by combining the parent directory, new file name, and original file extension
    new_file_path = os.path.join(parent_directory, new_file_name + os.path.splitext(filename)[1])

    # Copy the input file to the new file path
    shutil.copy2(input_file_path, new_file_path)


def copy_sheet_and_insert(source_filename, dest_filename, source_sheet_number, insert_after_sheet_number):
    # Load the workbooks
    source_wb = xl.load_workbook(source_filename)
    dest_wb = xl.load_workbook(dest_filename)

    # Get source and destination sheets
    src_sheet = source_wb.worksheets[source_sheet_number - 1]

    # Create a new sheet in the destination workbook
    new_sheet_index = insert_after_sheet_number + 1
    new_sheet = dest_wb.create_sheet(title=src_sheet.title, index=new_sheet_index)

    # Copy formulas, formatting, and styles
    for style_rows in src_sheet.iter_rows(min_row=1, max_row=src_sheet.max_row, min_col=1, max_col=src_sheet.max_column):
        for style_cells in style_rows:
            new_cells = new_sheet[style_cells.coordinate]
            new_cells.value = style_cells.value
            new_cells.font = copy(style_cells.font)
            new_cells.border = copy(style_cells.border)
            new_cells.fill = copy(style_cells.fill)
            new_cells.number_format = copy(style_cells.number_format)
            new_cells.protection = copy(style_cells.protection)
            new_cells.alignment = copy(style_cells.alignment)

    # Adjust column widths
    for col_widths in range(1, src_sheet.max_column + 1):
        col_ltr = xl.utils.get_column_letter(col_widths)
        new_sheet.column_dimensions[col_ltr].width = src_sheet.column_dimensions[col_ltr].width

    # Save changes to the destination workbook
    dest_wb.save(dest_filename)


def subtract_days_from_date(date_str, days):
    # Convert date string to datetime object
    date_object = datetime.strptime(date_str, '%m/%d/%Y')

    # Subtract the specified number of days
    new_date = date_object - timedelta(days=days)

    # Format the new date as mm/dd/yyyy
    new_date_str = new_date.strftime('%m%d%Y')

    return new_date_str


def dispatch():
    try:
        from win32com import client
        app = client.gencache.EnsureDispatch('Excel.Application')
    except AttributeError:
        # Corner case dependencies.
        import os
        import re
        import sys
        import shutil
        # Remove cache and try again.
        module_list = [m.__name__ for m in sys.modules.values()]
        for module in module_list:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        from win32com import client
        app = client.gencache.EnsureDispatch('Excel.Application')
    return app


dispatch()
kill_excel()

folder_path = 'Raw Payroll Invoices'
curr_dir = getcwd()
chdir(folder_path)

# Creating a separate copy of the raw Excel sheet as to preserve the original raw sheet.
temp_path = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\{folder_path}\temp.xlsx'

try:
    os.remove(temp_path)
except PermissionError:
    pass
except FileNotFoundError:
    pass

invoice_list = sorted(glob('*.xlsx'), key=os.path.getmtime)

for invoice in invoice_list:

    if invoice == 'temp.xlsx':
        try:
            os.remove(fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\{folder_path}\{invoice}')
        except FileNotFoundError:
            pass

    invoice_path = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\{folder_path}\{invoice}'
    wb = xl.load_workbook(invoice_path)

    source = wb.active
    target = wb.copy_worksheet(source)
    ws = wb.worksheets[1]
    ws1 = wb.worksheets[0]

    date = re.sub(r'\W+', '', str(ws.cell(row=2, column=2).value))
    date = date[:8]
    date = date[:4] + date[4:]
    date_dash = date[:2] + '-' + date[2:4] + '-' + date[4:]

    invoice_type = ws.cell(row=3, column=2).value

    ws.delete_rows(1, 4)


    temp_list = []

    # Checks if Cost Number has a value, if not then notify the user to review invoice.
    for i in range(6, ws.max_row - 1):
        if ws.cell(row=i, column=2).value is None or ws.cell(row=i, column=2).value == "":
            print(f'Employee Name is missing on row: {i}')
            temp_list.append(i)
        if ws.cell(row=i, column=3).value is None or ws.cell(row=i, column=3).value == "" or len(ws.cell(row=i, column=3).value) != 6 or "P74" in ws.cell(row=i, column=3).value:
            print(f'Cost Number is missing or incorrect, rename "{ws.cell(row=i, column=3).value}" on row: {i}')
            temp_list.append(i)
        if ws.cell(row=i, column=4).value is None or ws.cell(row=i, column=4).value == "":
            print(f'Cost Number Description is missing for "{ws.cell(row=i, column=4).value}" on row: {i}')

    if len(temp_list) != 0:
        sys.exit(1)
    else:
        pass

    # Add extra rows if invoice is missing "INV" and "INV-CHULA" rows.
    if invoice_type == 'P74':
        trunc_inv_list = []

        for i in range(6, ws.max_row - 1):
            trunc_inv_list.append(ws.cell(row=i, column=3).value[:4])

        trunc_string = ['SAFE', 'MELM']

        for i in trunc_string:
            if i not in trunc_inv_list:
                ws.insert_rows(ws.max_row - 1)
                ws.cell(row=ws.max_row - 2, column=3).value = i + '10'
                ws.cell(row=ws.max_row - 2, column=2).value = 'NA'
                ws.cell(row=ws.max_row - 2, column=4).value = 'NA'
    else:
        pass

    # Adds column for D3 column if the invoice doesn't have it.
    if locate("ADJ D3-UNION IN FEE") is None:
        ws.insert_cols(locate("ADJ 2-UNION")[1] + 1, 1)
        ws.cell(row=1, column=(locate("ADJ 2-UNION")[1] + 1)).value = "ADJ D3-UNION IN FEE"
        set_bold(f'A1:{get_column_letter(ws.max_column)}1')
    else:
        pass


    employer_adj = ws.cell(row=2, column=locate("Employer Adjustment Amount")[1]).value
    product_fee = ws.cell(row=2, column=locate("Ancillary Product Fee")[1]).value
    employee_fee_row = locate("Employee Ancillary Fee Amount")[1]

    sum_employee_fee = 0.00
    for i in range(2, ws.max_row-1):
        if ws.cell(row=i, column=employee_fee_row).value is None or ws.cell(row=i, column=employee_fee_row).value == '':
            sum_employee_fee += 0.00
        else:
            sum_employee_fee += float(ws.cell(row=i, column=employee_fee_row).value)

    sum_employee_fee = round(sum_employee_fee, 2)


    # Creates the code list explicitly (as opposed to Excel formulae) from "Cost Number#" column.
    code_list = []
    for i in range(2, ws.max_row + 1):
        code_list.append(ws.cell(row=i, column=3).value)


    # Exception handle to catch the wrong input format, if column is not "Cost Number#".
    try:
        if len(code_list[0]) == 6:
            code_list[:] = (code[:-2] for code in code_list if code is not None)
        else:
            sys.exit(1)
    except TypeError:
        print("Check if Cost Number column is in the correct format")
        sys.exit(1)

    ws.insert_cols(1, 1)
    for i, code in enumerate(code_list):
        ws.cell(row=i + 2, column=1).value = code
    ws.cell(row=1, column=1).value = "Code"
    ws.cell(row=1, column=2).value = "Check Date"

    title = f'PR-{date} {invoice_type}'
    ws.title = title
    ws1.title = f'{title} RAW'

    wb.save(temp_path)
    wb.close()

    #  Since Pandas deletes the other sheets from the Excel, I have to create another temporary Excel file, copy the sheet
    #  from the original Excel to the new one. Do the Pandas sorting in the other Excel File, then copy the new Excel sheet
    #  into the old Excel file. Then delete the original Excel sheet and rename the newly copied Excel sheet to the old name.
    temp_path1 = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\{folder_path}\temp1.xlsx'
    wb = xl.load_workbook(temp_path, data_only=True)
    ws = wb.worksheets[0]
    wb1 = xl.load_workbook(temp_path, data_only=True)
    ws1 = wb1.worksheets[0]

    for row in ws:
        for cell in row:
            try:
                ws1[cell.coordinate].value = cell.value
            except AttributeError:
                pass

    sheets_list = wb1.sheetnames

    del_list = []
    for i in sheets_list:
        if i != title:
            del_list.append(i)
        else:
            pass
    for i in del_list:
        del wb1[f"{i}"]

    wb1.save(temp_path1)
    wb1.close()

    wb1 = xl.load_workbook(temp_path1, data_only=True)
    ws1 = wb1.worksheets[0]

    # Pandas save automatically unlike openpyxl, which requires calling wb.save() to save. Pandas module sorts a column's
    # rows alphabetically or numerically. The column's row-first character value has to begin with letters OR numbers,
    # rows cannot contain both. Had to create Code column first, so it can be sorted by Code first, then Name column.
    df = pd.DataFrame(pd.read_excel(temp_path1, sheet_name=0))
    df.sort_values(["Code", "Name"], ascending=[True, True]).to_excel(temp_path1)
    del df

    wb1 = xl.load_workbook(temp_path1, data_only=True)
    ws1 = wb1.worksheets[0]

    # Pandas sort creates an extra column that shows which row moved, which I delete.
    ws1.delete_cols(1, 1)



    wb1.save(temp_path1)
    wb1.close()

    del wb[title]
    ws = wb.create_sheet(title, 0)

    for row in ws1:
        for cell in row:
            try:
                ws[cell.coordinate].value = cell.value
            except AttributeError:
                pass

    wb.save(temp_path)
    os.remove(temp_path1)


    # Pandas sorting complete.
    wb = xl.load_workbook(temp_path)
    ws = wb.worksheets[0]

    # Deletes unnecessary columns and rows from the Excel sheet.
    ws.delete_rows(locate("Grand Totals")[0], 2)
    ws.insert_cols(locate("Social Security Amount")[1], 1)
    ws.insert_cols(locate("Estimated Service Fee Amount")[1], 1)
    ws.insert_cols(locate("Sales & Use Tax")[1], 1)
    ws.insert_cols(locate("ADJ 78-DENTAL-POST TAX")[1], 1)
    ws.insert_cols(locate("Total Amount")[1], 2)
    ws.insert_cols(locate("ADJ 39-REIMBURSEMENT")[1], 1)

    write_header_list = [(locate("Gross Wages")[1] + 1, "ADJ GROSS WAGES"), (locate("SUI Amount")[1] + 1, "PR TAXES"),
                         (locate("WC Amount")[1] + 1, "DELIVERY"), (locate("ADJ NHF-NEW HIRE FEE")[1] + 1, "TOTAL WC"),
                         (locate("ADJ D3-UNION IN FEE")[1] + 1, "INS"), (locate("ADJ D3-UNION IN FEE")[1] + 2, "TOTAL AMT")]

    # Writes in the title or header for columns that don't have headers.
    for i in write_header_list:
        ws.cell(row=1, column=i[0]).value = i[1]

    corp_list = []

    # Finds the CORP rows.
    for i in range(0, ws.max_row):
        corp_list.append(ws.cell(row=i + 1, column=1).value)
    corp_listRows = []
    for i, ReferenceText in enumerate(corp_list):
        substring = ['CORP', 'OCOR', '0COR']
        try:
            if any(x in ReferenceText for x in substring):
                corp_listRows.append(i + 1)
        except TypeError:
            pass

    # Using the reverse of the rows makes it, so when moving the rows, the row number for the previous rows does not change.
    reverse_corp_list = []
    for i in reversed(corp_listRows):
        reverse_corp_list.append(i)


    # It copies the first reverse corp (or the last corp row) into the bottom row first then
    # copies the rest up, to make the corp row ordered alphabetical from the top.
    max_col = get_column_letter(locate("ADJ Subtotal")[1])
    for i, row in enumerate(reverse_corp_list):
        if i == 0:
            ws.move_range(f'A{row}:{max_col}{row}', rows=ws.max_row + 4 - row + len(corp_listRows) - 1, cols=0)
        else:
            ws.move_range(f'A{row}:{max_col}{row}', rows=ws.max_row-i-row, cols=0)
        ws.delete_rows(row, 1)


    # Finds the length of rows of FIELD data or non-CORP rows.
    field_total_idx = []
    i = -1
    r = locate("Cost Number")[1]
    while True:
        i += 1
        field_total_idx.append(ws.cell(row=i + 2, column=r).value)
        if None in field_total_idx:
            break

    # Total of field rows, the ending row number field_row + 1, since the title row.
    field_total_idx = list(filter(None, field_total_idx))
    field_row = len(field_total_idx)
    corp_row = len(corp_listRows) + field_row + 4

    # Range of columns with just data, no formulae.
    # Total row range of data from FIELD and CORP.
    column_range = [*range(locate("Gross Wages")[1], locate("Total Amount")[1] + 1), *range(locate("ADJ 39-REIMBURSEMENT")[1], locate("ADJ Subtotal")[1] + 1)]
    row_range = [*range(2, field_row + 2), *range(field_row + 5, corp_row + 1)]


    wb.save(temp_path)


    a1 = col_letter("Social Security Amount")
    a2 = col_letter("Medicare Amount")
    a3 = col_letter("FUTA Amount")
    a4 = col_letter("SUI Amount")

    b1 = col_letter("Estimated Service Fee Amount")
    b2 = col_letter("WC Amount")
    b3 = col_letter("DELIVERY")
    b4 = col_letter("Sales & Use Tax")
    b5 = col_letter("Refund Delivery Fees for Misrouted Check")
    b6 = col_letter("ADJ DTF-DRUG TESTING FEE")
    b7 = col_letter("ADJ NHF-NEW HIRE FEE")

    c1 = col_letter("ADJ 78-DENTAL-POST TAX")
    c2 = col_letter("ADJ B1-S CORP HEALTH")
    c3 = col_letter("ADJ B3-DPINS")
    c4 = col_letter("ADJ 31-MEDICAL")
    c5 = col_letter("ADJ 32-MEDICAL")
    c6 = col_letter("ADJ 33-TS DENTAL")
    c7 = col_letter("ADJ 34-VISION")
    c8 = col_letter("ADJ 79-VISION-POST TAX")
    c9 = col_letter("ADJ 2-UNION")
    c10 = col_letter("ADJ D3-UNION IN FEE")

    d1 = col_letter("Gross Wages")
    d2 = col_letter("PR TAXES")
    d3 = col_letter("TOTAL WC")
    d4 = col_letter("INS")

    e0 = col_letter("TOTAL AMT")
    e1 = col_letter("Total Amount")
    e2 = col_letter("ADJ H-MAJ MED POSTTAX")
    e3 = col_letter("ADJ 6-PREPAY")
    e4 = col_letter("ADJ D1-ROTH IRA")
    e5 = col_letter("ADJ Subtotal")
    e6 = col_letter("ADJ 39-REIMBURSEMENT")
    e7 = col_letter("ADJ Z-INS REIMBURSE")
    e8 = col_letter("ADJ GROSS WAGES")

    h = locate("ADJ GROSS WAGES")[1]
    w = locate("PR TAXES")[1]
    x = locate("TOTAL WC")[1]
    y = locate("INS")[1]
    z = locate("TOTAL AMT")[1]

    # Writes the Excel column sums of "ADJ GROSS WAGES, TOTAL Taxes", "Total WC", "INS", and "TOTAL".
    for i in row_range:
        ws.cell(row=i, column=h).value = f'={d1}{i}+{e6}{i}'
        ws.cell(row=i, column=w).value = f'={a1}{i}+{a2}{i}+{a3}{i}+{a4}{i}'
        ws.cell(row=i, column=x).value = f'={b1}{i}+{b2}{i}+{b3}{i}+{b4}{i}+{b5}{i}+{b6}{i}+{b7}{i}'
        ws.cell(row=i, column=y).value = f'={c1}{i}+{c2}{i}+{c3}{i}+{c4}{i}+{c5}{i}+{c6}{i}+{c7}{i}+{e7}{i}+{c8}{i}+{c9}{i}+{c10}{i}'
        ws.cell(row=i, column=z).value = f'={e8}{i}+{d2}{i}+{d3}{i}+{d4}{i}'

    for i in column_range:
        # Writes the sum total of FIELD and also the percentage and also changes the cell type format.
        coordinate = re.sub(r'[^a-zA-Z]', '', ws.cell(row=1, column=i).coordinate)
        ws.cell(row=field_row + 2, column=i).value = f'=SUM({coordinate}2:{coordinate}{field_row + 1})'
        ws[f'{coordinate}{field_row + 2}'].number_format = '$#,##0.00_);($#,##0.00)'
        ws.cell(row=field_row + 3, column=i).value = f'={coordinate}{field_row + 2}/${e8}${field_row + 2}'
        ws[f'{coordinate}{field_row + 3}'].number_format = '0.00%'

        # If there is no "Corp" rows, then don't write the sums.
        if locate("CORP", ws, 1):
            # Writes the sum total of CORP and also the percentage.
            ws.cell(row=corp_row + 1, column=i).value = f'=SUM({coordinate}{field_row + 5}:{coordinate}{corp_row})'
            ws[f'{coordinate}{corp_row + 1}'].number_format = '$#,##0.00_);($#,##0.00)'
            ws.cell(row=corp_row + 2, column=i).value = f'={coordinate}{corp_row + 1}/${e8}${corp_row + 1}'
            ws[f'{coordinate}{corp_row + 2}'].number_format = '0.00%'

            # Writes the total of "FIELD" and "CORP" at the bottom and their final sum.
            ws.cell(row=corp_row + 4, column=i).value = f'={coordinate}{field_row + 2}'
            ws.cell(row=corp_row + 5, column=i).value = f'={coordinate}{corp_row + 1}'
            ws.cell(row=corp_row + 6, column=i).value = f'={coordinate}{field_row + 2}+{coordinate}{corp_row + 1}'
        else:
            ws.cell(row=corp_row + 4, column=i).value = f'={coordinate}{field_row + 2}'
            ws.cell(row=corp_row + 6, column=i).value = f'={coordinate}{field_row + 2}'


        for j in range(4, 12):
            ws[f'{coordinate}{corp_row + j}'].number_format = '$#,##0.00_);($#,##0.00)'

        # Changes type the cell format of data with decimal and thousands separator, e.g., 1,000.00.
        for j in range(2, field_row + 2):
            ws[f'{coordinate}{j}'].number_format = '$#,##0.00_);($#,##0.00)'
        for k in range(field_row + 5, corp_row + 1):
            ws[f'{coordinate}{k}'].number_format = '$#,##0.00_);($#,##0.00)'

    ws.cell(row=corp_row + 4, column=3).value = "FIELD"
    ws.cell(row=corp_row + 5, column=3).value = "CORP"
    ws.cell(row=corp_row + 6, column=3).value = "TOTAL"

    j = locate("TOTAL AMT")[1] + 1
    ws.cell(row=corp_row + 7, column=j).value = "ADP"
    ws.cell(row=corp_row + 8, column=j).value = "MAJ MED"
    ws.cell(row=corp_row + 9, column=j).value = "PREPAY"
    ws.cell(row=corp_row + 10, column=j).value = "IRA"
    ws.cell(row=corp_row + 11, column=j).value = "DIFFERENCE"

    # Writes the "DELIVERY".
    v = locate("DELIVERY")[1]
    q = locate("ADJ DTF-DRUG TESTING FEE")[1]
    t = locate("ADJ NHF-NEW HIRE FEE")[1]


    # Changes to delivery fee
    for i in range(2, field_row + 2):
        if invoice_type == "RD4":
            ws.cell(row=i, column=v).value = f'={d1}{i}/${d1}${field_row + 2}*({sum_employee_fee}+{product_fee}+{product_fee})'
        else:
            ws.cell(row=i, column=v).value = f'={d1}{i}/${d1}${field_row + 2}*({employer_adj}+{sum_employee_fee})'


    for i in range(field_row + 5, corp_row + 1):
        ws.cell(row=i, column=v).value = 0
    for i in row_range:
        if ws.cell(row=i, column=q).value is None:
            ws.cell(row=i, column=q).value = 0
        if ws.cell(row=i, column=t).value is None:
            ws.cell(row=i, column=t).value = 0
        else:
            pass

    if locate("CORP", ws, 1):
        ws.cell(row=field_row + 5, column=v).value = product_fee
    else:
        pass

    # Changing date cell type format.
    for row in range(2, ws.max_row+1):
        ws[f'B{row}'].number_format = 'mm/dd/yy'


    # Calculation of the final difference in Excel.
    h = locate("TOTAL AMT")[1]
    p = locate("TOTAL")[0]
    ws.cell(row=p + 1, column=h).value = f'=-{col_letter("Total Amount")}{p}-{col_letter("DELIVERY")}{p}-{col_letter("Sales & Use Tax")}{p}'
    ws.cell(row=p + 2, column=h).value = f'={col_letter("ADJ H-MAJ MED POSTTAX")}{p}'
    ws.cell(row=p + 3, column=h).value = f'={col_letter("ADJ 6-PREPAY")}{p}'
    ws.cell(row=p + 4, column=h).value = f'={col_letter("ADJ D1-ROTH IRA")}{p}'
    ws.cell(row=p + 5, column=h).value = f'=SUM({col_letter("TOTAL AMT")}{p}:{col_letter("TOTAL AMT")}{p + 4})'

    # Manually change LMAG description
    try:
        L_MAG = locate("LMAG10")[1]
        for i in range(2, ws.max_row):
            if ws.cell(row=i, column=L_MAG).value == 'LMAG10':
                ws.cell(row=i, column=L_MAG + 1).value = 'Maintenance LMAG10'
    except (IndexError, TypeError) as error:
        pass

    set_font(f'A1:{max_col}{ws.max_row}')
    set_no_border(f'A1:{max_col}1')
    set_underline(f'A1:{max_col}1')
    set_alignment(f'A1:{max_col}1', 'left')
    set_bold(f'E{field_row + 2}:{max_col}{field_row + 2}')
    set_bold(f'E{corp_row + 1}:{max_col}{corp_row + 1}')
    set_bold(f'C{corp_row + 4}:C{corp_row + 4}')
    set_bold(f'C{corp_row + 5}:C{corp_row + 5}')
    set_bold(f'C{corp_row + 6}:{max_col}{corp_row + 6}')
    fill_color(f'A1:{max_col}1', 'C0C0C0')
    set_border(f'{col_letter("TOTAL AMT")}{corp_row + 11}:{col_letter("TOTAL AMT")}{corp_row + 11}', thin, empty, empty, double)

    red_font = Font(size=10, color='FF0000')
    ws.conditional_formatting.add(f'A1:{max_col}{ws.max_row}',
                                  CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=red_font))

    yellow_fill = PatternFill(bgColor="FFFF00")
    for i in row_range:
        ws.conditional_formatting.add(f'{max_col}{i}:{max_col}{i+1}',
                                      CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=yellow_fill))

    max_column_letter = xl.utils.get_column_letter(ws.max_column)

    manual_adjust(max_column_letter, [8, 13, 40, 21, 38, 13, 23, 25, 20, 20, 13, 13, 30, 15, 15, 20, 40, 28,
                                      28, 28, 28, 26, 16, 27, 25, 25, 20, 20, 20, 20, 16, 15, 22, 15, 18, 18, 10, 25, 23,
                                      25, 18, 18, 16, 24])

    # Freezes panes, so you can always see the title columns and code rows.
    ws.freeze_panes = 'B2'
    move_sheet(0, 1)


    wb.save(temp_path)
    wb.close()

    # ------------------------------------------------SUBTOTAL PART-----------------------------------------------------------------------------------------------------------------------------------------------------
    try:
        excel_obj = win32.gencache.EnsureDispatch('Excel.Application')
    except (AttributeError, TypeError):
        command = fr'C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe Remove-Item -path $env:LOCALAPPDATA\Temp\gen_py -recurse'
        try:
            subprocess.run(command, check=True)
        except subprocess.CalledProcessError as e:
            print("Error executing PowerShell command:", e)
            sys.exit(1)
        excel_obj = win32.Dispatch('Excel.Application')

    excel_obj.Visible = False

    subtotal_path = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\I-Workforce-Payroll-{date_dash}-{invoice_type}.xlsx'  # Change the user's directory for the Excel path for deployment.

    # Check if the Excel file already exists, removes the old one or creates a new one if it doesn't exist.
    try:
        os.remove(subtotal_path)
        pyxl_target_wb = xl.load_workbook(subtotal_path, data_only=True)
    except FileNotFoundError:
        pyxl_target_wb = xl.Workbook()
        pyxl_target_wb.save(subtotal_path)
        pyxl_target_wb = xl.load_workbook(subtotal_path, data_only=True)

    pywin_source_wb = excel_obj.Workbooks.Open(temp_path)
    pyxl_source_wb = xl.load_workbook(temp_path)
    pywin_target_wb = excel_obj.Workbooks.Open(subtotal_path)

    pywin_source_wb.Worksheets(pyxl_source_wb.sheetnames[1]).Copy(After=pywin_target_wb.Sheets(1))
    pywin_source_wb.Worksheets(pyxl_source_wb.sheetnames[0]).Copy(After=pywin_target_wb.Sheets(1))

    pywin_target_wb.Save()
    pywin_target_wb.Close()
    excel_obj.Quit()

    pyxl_target_wb = xl.load_workbook(subtotal_path, data_only=True)
    del pyxl_target_wb[pyxl_target_wb.sheetnames[0]]
    pyxl_target_wb.save(subtotal_path)
    pyxl_target_wb.close()


    subtotal_path = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\I-Workforce-Payroll-{date_dash}-{invoice_type}.xlsx'  # Change the user's directory for the Excel path for deployment.
    wb = xl.load_workbook(subtotal_path, data_only=True)   # Sets data in Excel to value only, no formulae.


    source = wb.active
    target = wb.copy_worksheet(source)

    ws = wb.worksheets[2]
    ws1 = wb.worksheets[1]

    # Initialize values to zero.
    num_hyder = 0
    num_adp = 0
    num_maj = 0
    num_prepay = 0
    num_ira = 0

    # Exception for RD4, since the Transfer sheet doesn't apply.
    try:
        num_hyder = abs(ws1.cell(row=locate('DIFFERENCE', ws1)[0] - 10, column=locate('DIFFERENCE', ws1)[1] - 1).value)
        num_adp = abs(ws1.cell(row=locate('DIFFERENCE', ws1)[0] - 4, column=locate('DIFFERENCE', ws1)[1] - 1).value)
        num_maj = abs(ws1.cell(row=locate('DIFFERENCE', ws1)[0] - 3, column=locate('DIFFERENCE', ws1)[1] - 1).value)
        num_prepay = abs(ws1.cell(row=locate('DIFFERENCE', ws1)[0] - 2, column=locate('DIFFERENCE', ws1)[1] - 1).value)
        num_ira = abs(ws1.cell(row=locate('DIFFERENCE', ws1)[0] - 1, column=locate('DIFFERENCE', ws1)[1] - 1).value)
    except TypeError:
        pass

    raw_date = re.sub(r'\W+', '', str(ws.cell(row=2, column=2).value))
    trunc_raw_date = raw_date[:8]
    date = trunc_raw_date[4:] + trunc_raw_date[:4]
    year = trunc_raw_date[:4]
    month_day = trunc_raw_date[4:]
    month = month_day[:2]
    day = month_day[2:]
    date_slash = date[:2] + '/' + date[2:4] + '/' + date[4:]

    date_time_str = f'{year}-{month}-{day}'
    date_time = datetime.fromisoformat(date_time_str)

    # Calculates the end date by doing "date" addition and subtraction.
    end_date_dt = date_time - timedelta(6)
    end_date = end_date_dt.strftime("%m%d%Y")

    title = f'PR-{date} P74'
    ws1.title = title

    try:
        del_col(["Code", "TOTAL AMT"])
        del_corp()
    except (IndexError, TypeError) as error:
        pass
    length()

    ws.delete_rows(length.f_row + 2, 100)

    #  The dictionary method below finds the subtotal of each code.
    code_List = []
    total_List = []

    for i in range(2, ws.max_row + 1):
        code_List.append(ws.cell(row=i, column=1).value)
        total_List.append(ws.cell(row=i, column=2).value)

    code_total_Tuple = list(tuple(zip(code_List, total_List)))

    code_total_Dict = {}
    # Create a dictionary with unique Codes and the associated Total values.
    for dict_key, dict_values in code_total_Tuple:
        code_total_Dict.setdefault(dict_key, []).append(dict_values)

    code_keys = []
    for key, value in code_total_Dict.items():
        code_keys.append(key)

    dictionary_List = []
    for i in range(0, len(code_total_Dict)):
        dictionary_List.append(i)

    ws.delete_rows(2, 10000)

    # Writes the dictionary key and values into the two columns.
    for i in range(0, len(code_total_Dict)):
        ws.cell(row=i + 2, column=1).value = str(list(code_total_Dict)[i]) + " Total"
        try:
            ws.cell(row=i + 2, column=2).value = sum(list(code_total_Dict.values())[i])
        except TypeError:
            print(f"Enter the correct {title} sheet")
            wb.save(subtotal_path)
            wb.close()
            sys.exit(1)


    ws.title = f'PR-{date} Transfer RAW'

    ws.insert_cols(1, 7)
    ws.insert_rows(1, 1)
    ws.cell(row=1, column=1).value = f"PR Check Date {date_slash}"
    ws.merge_cells('A1:J1')

    code_col = locate("Code")[1]
    code_list = []
    for i in range(3, ws.max_row + 2):
        code_list.append(ws.cell(row=i, column=code_col).value)

    code_list[:] = (code[:4] for code in code_list if code is not None)

    code_path = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\zOld\codes-current.xlsx'
    wb0 = xl.load_workbook(code_path)
    ws0 = wb0.worksheets[0]

    ADP_code = []

    for i in range(2, ws0.max_row+1):
        ADP_code.append(ws0.cell(row=i, column=3).value)

    for i, _ in enumerate(code_list):
        for j, _ in enumerate(ADP_code):
            if ADP_code[j] == code_list[i]:
                ws.cell(row=i + 3, column=1).value = ws0.cell(row=j + 2, column=1).value
                ws.cell(row=i + 3, column=2).value = ws0.cell(row=j + 2, column=4).value
                ws.cell(row=i + 3, column=3).value = ws0.cell(row=j + 2, column=5).value
                ws.cell(row=i + 3, column=4).value = ws0.cell(row=j + 2, column=6).value
                ws.cell(row=i + 3, column=5).value = ws0.cell(row=j + 2, column=7).value
                ws.cell(row=i + 3, column=6).value = ws0.cell(row=j + 2, column=2).value
                ws.cell(row=i + 3, column=7).value = ws0.cell(row=j + 2, column=3).value

    ws.cell(row=2, column=3).value = "Tax ID"
    ws.cell(row=2, column=4).value = "ABA"
    ws.cell(row=2, column=5).value = "Bank Account"
    ws.cell(row=2, column=6).value = "Yardi Code"
    ws.cell(row=2, column=7).value = "ADP - Code"
    ws.cell(row=2, column=10).value = "Invoice/Adjustment"
    ws.cell(row=ws.max_row + 1, column=8).value = "Grand Total"
    ws.cell(row=ws.max_row, column=9).value = f"=SUM(I3:I{ws.max_row - 1})"

    for i in range(3, len(code_list) + 5):
        ws[f'I{i}'].number_format = '$#,##0.00_);($#,##0.00)'

    set_font(f'A1:J{ws.max_row}')
    set_bold(f'A1:S2')
    set_underline(f'A2:S2')
    set_alignment(f'A1:S2', 'center')
    set_alignment(f'A3:H{ws.max_row}', 'left')
    no_fill(f'H2:I2')
    negative_red(f'A1:T{ws.max_row}')
    manual_adjust('J', [8, 34, 13, 13, 15, 13, 13, 24, 18, 24])

    ws.freeze_panes = 'B3'

    wb.save(subtotal_path)
    wb.close()


    wb = xl.load_workbook(subtotal_path)
    source = wb.worksheets[2]
    target = wb.copy_worksheet(source)
    wb.copy_worksheet(source)

    ws = wb.worksheets[3]
    ws.title = f'PR-{date} Transfer'

    ws.delete_rows(ws.max_row - 1, 2)

    wb.save(subtotal_path)


    ws = wb.worksheets[4]
    ws.title = f'PR-{date}-CBT-CSV'

    ws.unmerge_cells('A1:J1')
    ws.delete_rows(1, 2)
    ws.delete_rows(ws.max_row - 1, 2)

    ws.delete_cols(10, 1)
    ws.delete_cols(6, 3)
    ws.delete_cols(1, 1)

    del_cols = []
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=3).value != 122232109:

            del_cols.append(i)

    reverse_list = []
    for i in reversed(del_cols):
        reverse_list.append(i)

    for i in reverse_list:
        ws.delete_rows(i, 1)

    manual_adjust('E', [25, 15, 15, 15, 15])


    sort_excel_sheet(1, 1)

    wb.save(subtotal_path)
    wb.close()


    wb = xl.load_workbook(subtotal_path)
    source = wb.worksheets[1]
    target = wb.copy_worksheet(source)
    ws = wb.worksheets[5]

    # Changes the title of sheet to the date in the ADP Excel sheet.
    ws.title = f'PR-{date}-ETL RAW'
    try:
        del_corp()
    except (IndexError, TypeError) as error:
        pass
    del_col(["Code", "Check Date", "Name", "Cost Number", "Cost Number Description", "ADJ GROSS WAGES", "PR TAXES", "TOTAL WC", "INS", "TOTAL AMT"])
    length()

    ws.delete_rows(length.f_row + 2, 100)

    ws.insert_cols(10, 1)
    ws.insert_cols(9, 1)
    ws.insert_cols(8, 1)
    ws.insert_cols(7, 1)
    ws.insert_cols(4, 1)
    ws.insert_cols(2, 1)
    ws.insert_cols(1, 1)

    # Writes in the Yardi name, Yardi code, Department #, and Date/Description.
    for i in range(2, length.f_row + 2):
        for j, _ in enumerate(ADP_code):
            if ADP_code[j] == ws.cell(row=i, column=2).value:
                ws.cell(row=i, column=1).value = ws0.cell(row=j + 2, column=8).value
                ws.cell(row=i, column=3).value = ws0.cell(row=j + 2, column=2).value
                ws.cell(row=i, column=6).value = int(ws.cell(row=i, column=7).value[-2:])  # Needs to be integer for the comparison below.
                ws.cell(row=i, column=18).value = f'PPE{end_date} {ws.cell(row=i, column=8).value} / {ws.cell(row=i, column=5).value}'

        # Writes in the GL Codes.
        num = ws.cell(row=i, column=6).value
        if num == 10:
            ws.cell(row=i, column=10).value = '6510-0010'
        if num == 20:
            ws.cell(row=i, column=10).value = '6310-0000'
        if num == 30:
            ws.cell(row=i, column=10).value = '6330-0000'
        if num == 50:
            ws.cell(row=i, column=10).value = '6510-0020'
        if num == 60:
            ws.cell(row=i, column=10).value = '6510-0000'
        if num == 70:
            ws.cell(row=i, column=10).value = '6510-0030'
        if num == 80:
            ws.cell(row=i, column=10).value = '6330-0000'

        if num == 10 or num == 20 or num == 30 or num == 40 or num == 50 or num == 60 or num == 70 or num == 80:
            ws.cell(row=i, column=12).value = '6711-0000'
            ws.cell(row=i, column=14).value = '6722-0000'
            ws.cell(row=i, column=16).value = '6723-0000'

        prop_name = ws.cell(row=i, column=2).value

        # Changing the GL codes
        if num == 20 and (prop_name
                          == 'HOCH' or prop_name
                          == 'KNXG' or prop_name
                          == 'MAHO' or prop_name
                          == 'PAKI' or prop_name
                          == 'TOCO' or prop_name
                          == 'HDPA'):

            ws.cell(row=i, column=10).value = '6330-0000'

        if num == 60 and (prop_name
                          == 'HILL' or prop_name
                          == 'KMRI' or prop_name
                          == 'VAVI' or prop_name
                          == 'HOCH' or prop_name
                          == 'KNXG' or prop_name
                          == 'MAHO' or prop_name
                          == 'PAKI' or prop_name
                          == 'TOCO' or prop_name
                          == 'HDPA'):

            ws.cell(row=i, column=12).value = '6711-0011'
            ws.cell(row=i, column=14).value = '6722-0030'
            ws.cell(row=i, column=16).value = '6723-0030'

        if num == 10 and (prop_name
                          == 'HILL' or prop_name
                          == 'KMRI' or prop_name
                          == 'VAVI' or prop_name
                          == 'HOCH' or prop_name
                          == 'KNXG' or prop_name
                          == 'MAHO' or prop_name
                          == 'PAKI' or prop_name
                          == 'TOCO' or prop_name
                          == 'HDPA'):

            ws.cell(row=i, column=12).value = '6711-0010'
            ws.cell(row=i, column=14).value = '6722-0010'
            ws.cell(row=i, column=16).value = '6723-0010'

        if num == 50 and (prop_name
                          == 'HILL' or prop_name
                          == 'KMRI' or prop_name
                          == 'VAVI' or prop_name
                          == 'HOCH' or prop_name
                          == 'KNXG' or prop_name
                          == 'MAHO' or prop_name
                          == 'PAKI' or prop_name
                          == 'TOCO' or prop_name
                          == 'HDPA'):
            ws.cell(row=i, column=12).value = '6711-0020'
            ws.cell(row=i, column=14).value = '6722-0020'
            ws.cell(row=i, column=16).value = '6723-0020'


    set_font(f'A1:S{ws.max_row}')
    # Silver/Gray
    fill_color(f'B1:H1', 'C0C0C0')
    # Orange
    fill_color('I1:J1', 'FFC000')
    # Light Blue
    fill_color('K1:L1', 'DDEBF7')
    # Light Green
    fill_color('M1:N1', 'E2EFDA')
    # Light Orange
    fill_color('O1:P1', 'F8CBAD')
    set_alignment(f'A1:R{ws.max_row}', 'left')
    set_bold(f'A1:R1')
    set_underline(f'A1:R1')
    set_border(f'A1:R1', empty, empty, empty, thin)
    negative_red(f'A1:T{ws.max_row}')
    manual_adjust('R', [14, 8, 10, 12, 42, 6, 25, 38, 13, 15, 13, 14, 13, 12, 13, 13, 13, 80])
    ws.freeze_panes = 'B2'

    # Delete the non-formula sheet from workbook.
    del wb[wb.sheetnames[1]]

    wb.save(subtotal_path)
    wb.close()


    pywin_source_wb = excel_obj.Workbooks.Open(subtotal_path)
    pyxl_source_wb = xl.load_workbook(subtotal_path)
    pywin_target_wb = excel_obj.Workbooks.Open(subtotal_path)

    pywin_source_wb.Worksheets(pyxl_source_wb.sheetnames[4]).Copy(After=pywin_target_wb.Sheets(5))

    pywin_target_wb.Save()
    pywin_target_wb.Close()
    excel_obj.Quit()
    pyxl_source_wb.close()

    wb = xl.load_workbook(subtotal_path, data_only=True)
    ws = wb.worksheets[5]

    ws.title = f'PR-{date}-ETL-1'

    wb.save(subtotal_path)
    wb.close()


    wb = xl.load_workbook(subtotal_path)
    ws = wb.worksheets[5]

    error_check_list = []

    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value is None or ws.cell(row=i, column=2).value is None or ws.cell(row=i, column=7).value is None or len(ws.cell(row=i, column=2).value) != 4:
            error_check_list.append([ws.cell(row=i, column=7).value, f'Row:{i}'])

    first_column_values = []

    for i in range(2, ws.max_row + 1):
        first_column_values.append(ws.cell(row=i, column=1).value)

    # If there is an error, check if all these codes below are in the codes-current sheet.
    order_list = ['Yardi', 'Yardi-INV', 'Yardi-Inv-CV', 'Yardi-Western', 'Yardi-6911', 'Yardi-ADJ', 'Yardi-HACLA', 'Owner', 'Sage']

    # Sort the rows based on the first column values and the order list
    try:
        sorted_rows = sorted(range(len(first_column_values)), key=lambda b_1: order_list.index(first_column_values[b_1]))
    except ValueError:
        print(f"The program tried to sort the rows but failed.\n"
              f"\nCheck if these rows and columns, {error_check_list} have the incorrect values.\n"
              f"\nIf there are any mistakes, go into the raw invoice sheet and correct them.\n")
        print("Also make sure that the 'codes-current.xlsx' file is correct in the L Drive Payroll folder, add all the correct new properties to match with the payroll invoice.")
        wb.save(subtotal_path)
        wb.close()
        sys.exit(1)

    # Create a new worksheet to store the sorted rows
    sorted_worksheet = wb.create_sheet(title="Sorted")

    # Copy the column headers to the new worksheet
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            sorted_worksheet[cell.coordinate].value = cell.value

    # Copy the sorted rows to the new worksheet
    row_count = 2
    for row_index in sorted_rows:
        for row in ws.iter_rows(min_row=row_index + 2, max_row=row_index + 2):
            for cell in row:
                sorted_worksheet[cell.column_letter + str(row_count)].value = cell.value
        row_count += 1

    wb.save(subtotal_path)

    # Add back the formula sheet to workbook.
    pywin_source_wb = excel_obj.Workbooks.Open(temp_path)
    pyxl_source_wb = xl.load_workbook(temp_path)
    pywin_target_wb = excel_obj.Workbooks.Open(subtotal_path)

    pywin_source_wb.Worksheets(pyxl_source_wb.sheetnames[1]).Copy(After=pywin_target_wb.Sheets(1))

    pywin_target_wb.Save()
    pywin_target_wb.Close()
    excel_obj.Quit()

    kill_excel()

    wb = xl.load_workbook(subtotal_path, data_only=True)
    ws = wb.worksheets[6]

    source_sheet = 'Sorted'
    destination_sheet = wb.sheetnames[6]

    source_worksheet = wb[source_sheet]
    destination_worksheet = wb[destination_sheet]

    # Clear the existing data in the destination worksheet
    destination_worksheet.delete_rows(2, destination_worksheet.max_row)

    # Iterate through the rows in the source worksheet, skipping the first row
    for row in source_worksheet.iter_rows(min_row=2, values_only=True):
        destination_worksheet.append(row)

    del wb['Sorted']

    set_font(f'A2:S{ws.max_row + 3}')
    set_alignment(f'A2:S{ws.max_row + 3}', 'left')
    negative_red(f'A2:T{ws.max_row + 3}')

    columns = ['I', 'K', 'M', 'O', 'Q']
    for i in columns:
        for j in range(2, ws.max_row + 3):
            ws[f'{i}{j}'].number_format = '$#,##0.00_);($#,##0.00)'

    for row in range(2, ws.max_row+1):
        ws[f'D{row}'].number_format = 'mm/dd/yy'

    wb.save(subtotal_path)
    wb.close()

    # Skips Hacla separate sheet if there are no Hacla rows.
    try:
        wb = xl.load_workbook(subtotal_path)
        source = wb.worksheets[6]
        target = wb.copy_worksheet(source)
        wb.copy_worksheet(source)

        ws = wb.worksheets[7]
        ws.title = f'PR-{date}-ETL-2'
        ws.delete_rows(locate('Yardi-HACLA')[0], 100)

        ws = wb.worksheets[8]
        ws.title = f'PR-{date}-HACLA'

        ws.delete_rows(2, locate('Yardi-HACLA')[0] - 2)

        hacla_col = []
        for i in range(1, ws.max_row):
            if ws.cell(row=i, column=1).value == 'Yardi-HACLA':
                hacla_col.append(ws.cell(row=i, column=1).value)

        ws.delete_rows(len(hacla_col) + 2, 100)

        col_num = locate('TOTAL AMT')[1]
        col_let = col_letter('TOTAL AMT')

        ws.cell(row=ws.max_row + 2, column=col_num).value = f"=SUM({col_let}2:{col_let}{ws.max_row})"
        ws[f'{col_let}{ws.max_row}'].number_format = '$#,##0.00_);($#,##0.00)'
        set_bold(f'{col_let}{ws.max_row}:{col_let}{ws.max_row}')
        set_alignment(f'{col_let}{ws.max_row}:{col_let}{ws.max_row}', 'left')

        wb.save(subtotal_path)
        wb.close()
    except (IndexError, TypeError) as error:
        pass

    try:
        ws = wb.worksheets[6]
        ws.insert_rows(locate('Yardi-6911')[0], 3)
        ws.insert_rows(locate('Yardi-ADJ')[0], 3)
        ws.insert_rows(locate('Yardi-HACLA')[0], 3)
        ws.insert_rows(locate('Owner')[0], 3)

        max_rows = []

        for i in range(1, ws.max_row):
            max_rows.append(ws.cell(row=i, column=1).value)

        reverse_list = []
        for i in reversed(max_rows):
            reverse_list.append(i)

        count = 0
        for i in reverse_list:
            if i is None:
                count += 1
            elif i is not None:
                break

        max_row_len = len(reverse_list) - count

        # Writing in the sums of the rows.
        ws.cell(row=locate('Yardi-6911')[0] - 3,
                column=col_num).value = f"=SUM({col_let}2:{col_let}{locate('Yardi-6911')[0] - 4})"
        ws.cell(row=locate('Yardi-ADJ')[0] - 3,
                column=col_num).value = f"=SUM({col_let}{locate('Yardi-6911')[0]}:{col_let}{locate('Yardi-ADJ')[0] - 4})"
        ws.cell(row=locate('Yardi-HACLA')[0] - 3,
                column=col_num).value = f"=SUM({col_let}{locate('Yardi-ADJ')[0]}:{col_let}{locate('Yardi-HACLA')[0] - 4})"
        ws.cell(row=locate('Owner')[0] - 3,
                column=col_num).value = f"=SUM({col_let}{locate('Yardi-HACLA')[0]}:{col_let}{locate('Owner')[0] - 4})"
        ws.cell(row=max_row_len + 1, column=col_num).value = f"=SUM({col_let}{locate('Owner')[0]}:{col_let}{max_row_len})"
        ws.cell(row=max_row_len + 3, column=col_num).value = f"=SUM({col_let}{locate('Yardi-6911')[0] - 3}+" \
                                                             f"{col_let}{locate('Yardi-ADJ')[0] - 3}+" \
                                                             f"{col_let}{locate('Yardi-HACLA')[0] - 3}+" \
                                                             f"{col_let}{locate('Owner')[0] - 3}+" \
                                                             f"{col_let}{max_row_len + 1})"

        set_font(f'A2:S{max_row_len + 3}')
        set_alignment(f'A2:S{max_row_len + 3}', 'left')

        bold_list = [f"{col_let}{locate('Yardi-6911')[0] - 3}:{col_let}{locate('Yardi-6911')[0] - 3}",
                     f"{col_let}{locate('Yardi-ADJ')[0] - 3}:{col_let}{locate('Yardi-ADJ')[0] - 3}",
                     f"{col_let}{locate('Yardi-HACLA')[0] - 3}:{col_let}{locate('Yardi-HACLA')[0] - 3}",
                     f"{col_let}{locate('Owner')[0] - 3}:{col_let}{locate('Owner')[0] - 3}",
                     f"{col_let}{max_row_len + 1}:{col_let}{max_row_len + 1}",
                     f"{col_let}{max_row_len + 3}:{col_let}{max_row_len + 3}"]

        for i in bold_list:
            set_bold(i)

        negative_red(f'A2:T{max_row_len + 3}')

        columns = ['I', 'K', 'M', 'O', 'Q']
        for i in columns:
            for j in range(2, max_row_len + 3):
                ws[f'{i}{j}'].number_format = '$#,##0.00_);($#,##0.00)'

        # Writing this sum here because before, it would change into a value, instead of the Excel formula.
        ws = wb.worksheets[4]

        ws.cell(row=ws.max_row + 2, column=5).value = f"=SUM(E1:E{ws.max_row})"
        ws[f'E{ws.max_row}'].number_format = '$#,##0.00_);($#,##0.00)'
        set_bold(f'E{ws.max_row}:E{ws.max_row}')

        wb.save(subtotal_path)
        wb.close()

    except (IndexError, TypeError) as error:
        pass

    # Exception for RD4, since the Transfer sheet and labor distribution doesn't apply.
    if invoice_type == 'P74':
        # Moved Transfer Excel down here so that the formulas would be preserved.
        wb = xl.load_workbook(subtotal_path)
        ws = wb.worksheets[3]

        try:
            ws.cell(row=locate("SUNB")[0], column=4).value = "INV-RM"
        except (IndexError, TypeError) as error:
            pass

        # Error is caused by None type in columns when trying to sort, this exception catches it and lists all the needed changes.
        try:
            sort_excel_sheet(3, 4)
        except TypeError:
            print('\nMake sure that the "codes-current.xlsx" file is correct in the L Drive Payroll folder,\n'
                  'Add all the new property information corresponding to the ADP codes listed below:\n')
            for i in range(3, ws.max_row + 1):
                if ws.cell(row=i, column=2).value is None:
                    print(ws.cell(row=i, column=8).value[:-6])
            print('\nAfter adding the new properties, save the codes-current Excel file and rerun the program again.')
            wb.save(subtotal_path)
            wb.close()
            sys.exit(1)

        # Creating a sort list in order to group each portion by added 3 empty rows between them when the "ABA" name changes.
        sorted_list = []
        for i in range(3, ws.max_row + 1):
            sorted_list.append(ws.cell(row=i, column=4).value)

        indexes = []
        previous_value = None
        for i, value in enumerate(sorted_list):
            if value != previous_value:
                indexes.append(i + 3)
                previous_value = value

        indexes.pop(0)


        for i in reversed(indexes):
            if (ws.cell(row=i, column=4).value == "INV-CHULA" or ws.cell(row=i, column=4).value == "INV-HACLA" or ws.cell(
                    row=i, column=4).value == "INV" or
                    ws.cell(row=i, column=4).value == "INV-RM"):
                ws.insert_rows(i, 3)


        # Change back those values since these groups are already grouped.
        try:
            ws.cell(row=locate("SUNB")[0], column=4).value = "INV"
        except (IndexError, TypeError) as error:
            pass

        # Added all the sums, formulas and formatting to the sheet.
        ws.cell(row=locate("INV")[0] - 3, column=8).value = "Total ADJ"
        ws.cell(row=locate("INV")[0] - 2, column=8).value = "Yardi Total"
        ws.cell(row=locate("INV-CHULA")[0] - 3, column=8).value = "Total Prop"
        ws.cell(row=locate("INV-HACLA")[0] - 3, column=8).value = "Total CV Invoice"
        ws.cell(row=locate("INV-RM")[0] - 3, column=8).value = "Total HACLA Invoice"
        ws.cell(row=ws.max_row + 1, column=8).value = "Total Owners"
        ws.cell(row=locate("Total Owners")[0] + 3, column=8).value = "ACH"
        ws.cell(row=locate("Total Owners")[0] + 4, column=8).value = "Property ACH"
        ws.cell(row=locate("Total Owners")[0] + 5, column=8).value = "Hyder"
        ws.cell(row=locate("Total Owners")[0] + 6, column=8).value = "Total ACH"
        ws.cell(row=locate("Total Owners")[0] + 8, column=8).value = "Total Properties"
        ws.cell(row=locate("Total Owners")[0] + 9, column=8).value = "Hyder"
        ws.cell(row=locate("Total Owners")[0] + 10, column=8).value = "Total Invoice Amount"
        ws.cell(row=locate("Total Owners")[0] + 12, column=8).value = "Reconciliation"
        ws.cell(row=locate("Total Owners")[0] + 13, column=8).value = "ADP - CBT"
        ws.cell(row=locate("Total Owners")[0] + 14, column=8).value = "Prepay"
        ws.cell(row=locate("Total Owners")[0] + 15, column=8).value = "Roth IRA"
        ws.cell(row=locate("Total Owners")[0] + 16, column=8).value = "MMP"
        ws.cell(row=locate("Total Owners")[0] + 17, column=8).value = "Total Invoice - ADP"
        ws.cell(row=locate("Total Owners")[0] + 17, column=10).value = "DIFFERENCE"
        ws.cell(row=locate("Total Owners")[0] + 13, column=10).value = "ADP"
        ws.cell(row=locate("Total Owners")[0] + 14, column=10).value = "PREPAY"
        ws.cell(row=locate("Total Owners")[0] + 15, column=10).value = "IRA"
        ws.cell(row=locate("Total Owners")[0] + 16, column=10).value = "MAJ MED"

        # Writing sums of each group, Total Yardi-ADJ, Yardi Total, Chula Vista Invoice, HACLA Invoice, Owners' Invoice.
        ws.cell(row=locate("INV")[0] - 3, column=9).value = f'=SUM(I3:I{locate("INV")[0] - 4})'
        ws.cell(row=locate("INV-CHULA")[0] - 3, column=9).value = f'=SUM(I{locate("INV")[0]}:I{locate("INV-CHULA")[0] - 4})'
        ws.cell(row=locate("INV-HACLA")[0] - 3,
                column=9).value = f'=SUM(I{locate("INV-CHULA")[0]}:I{locate("INV-HACLA")[0] - 4})'
        ws.cell(row=locate("INV-RM")[0] - 3, column=9).value = f'=SUM(I{locate("INV-HACLA")[0]}:I{locate("INV-RM")[0] - 4})'
        ws.cell(row=locate("Total Owners")[0],
                column=9).value = f'=SUM(I{locate("Total HACLA Invoice")[0] + 3}:I{locate("Total Owners")[0] - 1})'
        ws.cell(row=locate("INV")[0] - 2,
                column=9).value = f'=SUM(I{locate("INV")[0] - 3}+I{locate("INV-CHULA")[0] - 3}+I{locate("INV-HACLA")[0] - 3}+I{locate("INV-RM")[0] - 3})'

        # Writing the values and sums at the summary at the bottom.
        ws.cell(row=locate("Property ACH")[0], column=9).value = ws.cell(row=locate("Total ADJ")[0], column=9).value
        try:
            ws.cell(row=locate("Property ACH")[0] + 1, column=9).value = num_hyder
        except (IndexError, TypeError, NameError) as error:
            pass
        ws.cell(row=locate("Property ACH")[0] + 2,
                column=9).value = f'=SUM(I{locate("Property ACH")[0]}:I{locate("Property ACH")[0] + 1})'

        ws.cell(row=locate("Total ACH")[0] + 2,
                column=9).value = f'=SUM(I{locate("Yardi Total")[0]}+I{locate("Total Owners")[0]})'
        try:
            ws.cell(row=locate("Total ACH")[0] + 3, column=9).value = num_hyder
        except (IndexError, TypeError, NameError) as error:
            pass
        ws.cell(row=locate("Total ACH")[0] + 4,
                column=9).value = f'=SUM(I{locate("Total ACH")[0] + 2}:I{locate("Total ACH")[0] + 3})'

        ws.cell(row=locate("Reconciliation")[0] + 1, column=9).value = num_adp
        ws.cell(row=locate("Reconciliation")[0] + 2, column=9).value = num_prepay
        ws.cell(row=locate("Reconciliation")[0] + 3, column=9).value = num_ira
        ws.cell(row=locate("Reconciliation")[0] + 4, column=9).value = num_maj
        ws.cell(row=locate("Reconciliation")[0] + 5,
                column=9).value = f'=SUM(I{locate("Reconciliation")[0] + 1}:I{locate("Reconciliation")[0] + 4})'
        ws.cell(row=locate("Reconciliation")[0] + 5, column=11).value = f'=SUM(I{locate("Reconciliation")[0] - 2}-I{locate("Reconciliation")[0] + 5})'

        # Formatting and styling the sheet.
        for i in range(3, ws.max_row + 1):
            coordinate = re.sub(r'[^a-zA-Z]', '', ws.cell(row=i, column=9).coordinate)
            ws[f'{coordinate}{i}'].number_format = '$#,##0.00_);(#,##0.00)'

        ws[f'K{locate("DIFFERENCE")[0]}'].number_format = '$#,##0.00_);(#,##0.00)'

        negative_red(f'A3:J{ws.max_row + 1}')
        set_font(f'A3:K{ws.max_row + 1}')

        set_border(f'H{locate("Total ADJ")[0]}:I{locate("Total ADJ")[0]}', thin, empty, empty, empty)
        set_border(f'H{locate("Total ADJ")[0] + 1}:I{locate("Total ADJ")[0] + 1}', thin, empty, empty, double)
        set_bold(f'H{locate("Total ADJ")[0]}:I{locate("Total ADJ")[0] + 1}')

        set_border(f'H{locate("Total Prop")[0]}:I{locate("Total Prop")[0]}', thin, empty, empty, double)
        set_bold(f'H{locate("Total Prop")[0]}:I{locate("Total Prop")[0]}')

        set_border(f'H{locate("Total CV Invoice")[0]}:I{locate("Total CV Invoice")[0]}', thin, empty, empty, double)
        set_bold(f'H{locate("Total CV Invoice")[0]}:I{locate("Total CV Invoice")[0]}')

        set_border(f'H{locate("Total HACLA Invoice")[0]}:I{locate("Total HACLA Invoice")[0]}', thin, empty, empty, double)
        set_bold(f'H{locate("Total HACLA Invoice")[0]}:I{locate("Total HACLA Invoice")[0]}')

        set_border(f'H{locate("Total Owners")[0]}:I{locate("Total Owners")[0]}', thin, empty, empty, double)
        set_bold(f'H{locate("Total Owners")[0]}:I{locate("Total Owners")[0]}')

        set_border(f'H{locate("Total ACH")[0]}:I{locate("Total ACH")[0]}', thin, empty, empty, double)
        set_bold(f'H{locate("Total ACH")[0]}:I{locate("Total ACH")[0]}')

        set_border(f'H{locate("Total ACH")[0] + 4}:I{locate("Total ACH")[0] + 4}', thin, empty, empty, double)
        set_bold(f'H{locate("Total ACH")[0] + 4}:I{locate("Total ACH")[0] + 4}')

        set_border(f'H{locate("Total ACH")[0] + 11}:I{locate("Total ACH")[0] + 11}', thin, empty, empty, double)
        set_bold(f'H{locate("Total ACH")[0] + 11}:I{locate("Total ACH")[0] + 11}')

        set_bold(f'H{locate("ACH")[0]}:I{locate("ACH")[0]}')
        set_bold(f'H{locate("Reconciliation")[0]}:I{locate("Reconciliation")[0]}')

        set_border(f'K{locate("DIFFERENCE")[0]}:K{locate("DIFFERENCE")[0]}', thin, empty, empty, double)


        wb.save(subtotal_path)

        # Copying the two sheets from the temp files and pasting it back into the main template file.
        xl_obj = Dispatch('Excel.Application')
        xl.Visible = False

        wb = xl_obj.Workbooks.Open(subtotal_path)
        wb1 = xl_obj.Workbooks.Open(temp_path)

        ws1 = wb1.Worksheets(2)
        ws1.Copy(After=wb.Worksheets(1))

        wb.Save()
        wb1.Close()

        wb.Close()
        xl_obj.Quit()

        # Rename the sheet with no formulas into P74 Final.
        wb = xl.load_workbook(subtotal_path)

        # Have to change the second sheet first because the two sheets would have the same name.
        wb.worksheets[2].title += " Final"
        wb.worksheets[1].title = wb.worksheets[1].title[:-4]

        wb.save(subtotal_path)
        wb.close()

        # Remove the temp files in the Raw Payroll Invoices folder.
        os.remove(temp_path)

        # Hide columns for Final P74.
        col_ranges = [['A', 'E'], ['G', 'G'], ['W', 'W'], ['AH', 'AI'], ['AJ', 'AQ']]

        wb = load_workbook(subtotal_path)
        ws = wb.worksheets[2]

        ws.delete_cols(36, 9)
        ws.delete_cols(24, 10)
        ws.delete_cols(13, 10)
        ws.delete_cols(8, 4)
        ws.delete_cols(6, 1)

        # Delete Corps rows.
        ws.delete_rows(field_row + 2, corp_row - field_row + 1)


        subtotal_list = []

        for i in range(2, field_row + 2):
            subtotal_list.append([ws.cell(row=i, column=1).value, '', '', '', '',
                                  ws.cell(row=i, column=6).value,
                                  ws.cell(row=i, column=7).value, ws.cell(row=i, column=8).value,
                                  ws.cell(row=i, column=9).value, ws.cell(row=i, column=10).value])


        summed_subtotal_list = []
        current_first_element = None
        sum_last_five = [0] * 5

        for row in subtotal_list:
            first_element = row[0]
            if first_element != current_first_element:
                if current_first_element is not None:
                    # Append the 'Total' row
                    total_row = [current_first_element + ' Total'] + [''] * 4 + sum_last_five
                    summed_subtotal_list.append(total_row)
                    sum_last_five = [0] * 5
                current_first_element = first_element

            # Add the row values to the sums
            for i in range(-5, 0):
                sum_last_five[i] += row[i] if isinstance(row[i], (int, float)) else 0

            # Append the row itself
            summed_subtotal_list.append(row)

        # Append the last 'Total' row
        if current_first_element is not None:
            total_row = [current_first_element + ' Total'] + [''] * 4 + sum_last_five
            summed_subtotal_list.append(total_row)


        position_subtotal_list = []
        for i, j in enumerate(summed_subtotal_list):
            if len(j[0]) != 4:
                position_subtotal_list.append([j, i + 1])


        for position_rows in position_subtotal_list:
            ws.insert_rows(position_rows[1] + 1)
            for j, k in enumerate(position_rows[0]):
                ws.cell(row=position_rows[1] + 1, column=j + 1).value = k

        for j in range(2, ws.max_row - 8):
            for i in range(5, 11):
                coordinate = re.sub(r'[^a-zA-Z]', '', ws.cell(row=j, column=i).coordinate)
                ws[f'{coordinate}{j}'].number_format = '$#,##0.00_);($#,##0.00)'


        for position_rows in position_subtotal_list:
            fill_color(f'A{position_rows[1] + 1}:J{position_rows[1] + 1}', 'FFFF00')
            set_border(f'A{position_rows[1] + 1}:J{position_rows[1] + 1}', thin, empty, empty, double)
            set_bold(f'A{position_rows[1] + 1}:A{position_rows[1] + 1}')


        set_font(f'B2:K{ws.max_row - 8}')
        negative_red(f'A2:K{ws.max_row + 1}')

        manual_adjust("K", [12, 13, 40, 16, 38, 20, 18, 18, 18, 18])

        wb.save(subtotal_path)
        wb.close()

        wb = xl.load_workbook(subtotal_path)
        ws = wb.worksheets[1]

        # Deleting extra rows for "Inv" and "Inv-Chula". Cannot delete rows from sheets that have formulas since rows won't move.
        if ws.cell(row=locate('MELM', ws)[0], column=2).value is None:
            for i in [2, 3, 5, 6, 7, 8, 9]:
                wb = xl.load_workbook(subtotal_path)
                ws = wb.worksheets[i]

                delete_inv_rows = ['SAFE Total', 'SAFE', 'MELM Total', 'MELM']

                for j in delete_inv_rows:
                    try:
                        ws.delete_rows(locate(j, ws)[0])
                        wb.save(subtotal_path)
                        wb.close()
                    except IndexError:
                        pass

        print("Payroll file complete.")


        # ----------------------------------------------------Labor Distribution Part---------------------------------------------------------------------------------------------------------------

        labor_path = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\{date_dash} Labor Distribution.xlsx'
        input_file = fr'C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Payroll\zOld\Labor Distribution Template.xlsx'
        new_filename = fr'{date_dash} Labor Distribution'

        try:
            copy_excel_file(input_file, new_filename)
        except FileExistsError:
            pass

        wb = load_workbook(labor_path)
        ws = wb.worksheets[0]

        wb1 = xl.load_workbook(subtotal_path)
        ws1 = wb1.worksheets[2]

        # Copy formulas, formatting, and styles
        for _row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
            for cell in _row:
                new_cell = ws[cell.coordinate]
                new_cell.value = cell.value
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        # Adjust column widths
        for _col in range(1, ws1.max_column + 1):
            column_letter = xl.utils.get_column_letter(_col)
            ws.column_dimensions[column_letter].width = ws1.column_dimensions[column_letter].width

        wb.save(labor_path)
        wb.close()

        # Generate all the invoice sheets for each property that has a value.
        wb = xl.load_workbook(labor_path)
        ws = wb.worksheets[2]
        wb1 = xl.load_workbook(subtotal_path)
        ws1 = wb1.worksheets[3]
        ws0 = wb.worksheets[1]
        ws2 = wb.worksheets[0]

        # delete_inv_rows = ['SAFE Total', 'SAFE', 'MELM Total', 'MELM']
        #
        # for i in delete_inv_rows:
        #     if ws2.cell(row=locate(i, ws2)[0], column=2).value is None:
        #         ws2.delete_rows(locate(i, ws2)[0])



        prop_code_list = []
        for i in range(3, ws1.max_row - 1):
            prop_code_list.append(ws1.cell(row=i, column=1).value)

        prop_code_tuple = []
        for i in range(3, ws1.max_row - 1):
            prop_code_tuple.append((ws1.cell(row=i, column=1).value, ws1.cell(row=i, column=2).value))

        labour_code_list = []
        for i in range(9, locate("TOTAL:", ws0, 2, 2)[0] - 1):
            labour_code_list.append(ws0.cell(row=i, column=1).value)

        # Checks if properties in Payroll workbook are also in Labor template.
        missing_list = []

        for i in set(prop_code_list):
            if i not in labour_code_list:
                missing_list.append(i)

        if missing_list:
            print(f"\nCheck if the associated {missing_list} codes are in the Labor Distribution Template,"
                  f"\nif not, add rows by copy and inserting rows, renaming the code, 'XXXX20' to new code, and rerun program.")
            sys.exit(1)
        else:
            pass

        ws0.cell(row=2, column=3).value = date_slash
        ws0.cell(row=3, column=3).value = f'PPE-{subtract_days_from_date(date_slash, 6)}'
        ws0.cell(row=4, column=3).value = date_slash
        ws0.cell(row=5, column=3).value = date_slash


        ws0.cell(row=locate('CORP', ws0, 1)[0], column=3).value = ws2.cell(row=ws2.max_row - 6, column=6).value
        ws0.cell(row=locate('CORP', ws0, 1)[0], column=10).value = ws2.cell(row=ws2.max_row - 6, column=7).value
        ws0.cell(row=locate('CORP', ws0, 1)[0], column=14).value = ws2.cell(row=ws2.max_row - 6, column=8).value
        ws0.cell(row=locate('CORP', ws0, 1)[0], column=18).value = ws2.cell(row=ws2.max_row - 6, column=9).value
        ws0.cell(row=locate('CORP', ws0, 1)[0], column=23).value = ws2.cell(row=ws2.max_row - 6, column=10).value

        for i, j in enumerate(prop_code_tuple):
            wb.copy_worksheet(wb.worksheets[2])
            ws = wb.worksheets[3 + i]


            ws.cell(row=1, column=1).value = j[1]

            # Writing the job position values for each invoice.
            ws.cell(row=13,
                    column=6).value = f'=SUMIFS(PR!$F$2:$F$993,PR!$D$2:$D$993,"{j[0]}30",PR!$D$2:$D$993,"{j[0]}30")'
            ws.cell(row=15,
                    column=6).value = f'=SUMIFS(PR!$F$2:$F$993,PR!$D$2:$D$993,"{j[0]}20",PR!$D$2:$D$993,"{j[0]}20")'
            ws.cell(row=17,
                    column=6).value = f'=SUMIFS(PR!$F$2:$F$993,PR!$D$2:$D$993,"{j[0]}10",PR!$D$2:$D$993,"{j[0]}10")'
            ws.cell(row=19,
                    column=6).value = f'=SUMIFS(PR!$F$2:$F$993,PR!$D$2:$D$993,"{j[0]}50",PR!$D$2:$D$993,"{j[0]}50")'
            ws.cell(row=21,
                    column=6).value = f'=SUMIFS(PR!$F$2:$F$993,PR!$D$2:$D$993,"{j[0]}70",PR!$D$2:$D$993,"{j[0]}70")'
            ws.cell(row=23,
                    column=6).value = f'=SUMIFS(PR!$F$2:$F$993,PR!$D$2:$D$993,"{j[0]}60",PR!$D$2:$D$993,"{j[0]}60")'

            try:
                adp_row = locate(j[0], ws0, 1)[0]
            except TypeError:
                print("Error in ADP Code List. Check in subtotal or labor sheet.")
                sys.exit(1)

            if ws0.cell(row=adp_row, column=9).value is None:
                ws.cell(row=25, column=6).value = 0.00
            else:
                ws.cell(row=25, column=6).value = ws0.cell(row=adp_row, column=9).value

            ws.cell(row=27, column=6).value = f'=MASTERPPEBIL!J{adp_row}'
            ws.cell(row=29, column=6).value = f'=MASTERPPEBIL!N{adp_row}'
            ws.cell(row=31, column=6).value = f'=MASTERPPEBIL!R{adp_row}'

            ws.title = j[1]

        wb.save(labor_path)
        wb.close()

        # Writing Hyder Invoice sheet
        wb = xl.load_workbook(labor_path)
        ws = wb.worksheets[2]

        ws.cell(row=1, column=1).value = 'Hyder'

        ws.cell(row=13, column=6).value = ws0.cell(row=locate('CORP', ws0)[0], column=3).value
        ws.cell(row=27, column=6).value = ws0.cell(row=locate('CORP', ws0)[0], column=10).value
        ws.cell(row=29, column=6).value = ws0.cell(row=locate('CORP', ws0)[0], column=14).value
        ws.cell(row=31, column=6).value = ws0.cell(row=locate('CORP', ws0)[0], column=18).value

        ws.delete_rows(15, 12)
        ws.cell(row=22, column=6).value = f'=SUM(F13:F21)'
        ws.title = 'Hyder'

        wb.save(labor_path)
        wb.close()

        kill_excel()
        print("Labor Distribution file complete.")

    else:
        os.remove(temp_path)

print("------- Program finished running in %s seconds. -------" % round((time.time() - start_time), 2))
